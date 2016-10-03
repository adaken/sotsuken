# coding: utf-8

import numpy as np
import matplotlib.pyplot as plt
import openpyxl as px


if __name__ == '__main__':

    xlsx_path = "E:\work\gps_hirano.xlsx"

    # xlsx読み込み
    wb = px.load_workbook(filename=xlsx_path, read_only=True)
    ws = wb['Sheet1']

    """
    ヘッダを含まずにrow_offset+1行目から読み込む
    """
    print "select columns..."
    acs = [] # 加速度のリスト
    col_letters = ('M') # Excelの列のレター
    row_offset = 2 # 読み込みを開始する行
    max_row = ws.max_row # 最後の行
    # 列を抜き出す
    for letter in col_letters:
        col = [] # 列
        for cell in ws['%s%d:%s' % (letter, row_offset, max_row)]:
            for data in cell:
                acs.append(data.value)

    n = 128
    fft= np.fft.fft(acs[0:n])[0:n/2]
    print fft
    
    afft = abs(fft)
    #for i in xrange(0, len(acs), n):
        #fft = np.fft.fft(acs[i:])[0:n/2]
    x = np.arange(acs[0:n])
    #y = np.fft.fft(np.sin(x))
    freq = np.fft.fft(x.shape[-1])
    plt.plot(freq, fft.real, freq, fft.imag)
    plt.show()