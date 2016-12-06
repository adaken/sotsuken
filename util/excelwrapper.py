# coding: utf-8

import openpyxl as px
import re

class ExcelWrapper(object):
    """
    Excelファイルを扱うラッパークラスです
    openpyxlを使ってExcelファイルを操作します
    主にファイル読み込みなどを抽象化してあります
    より深いことを行い場合はopenpyxlを直接使おう

    基本的にシート単位で操作するので、
    インスタンスを作った後、get_sheet()を使ってシートのオブジェクトを入手してください
    """

    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def __init__(self, filename):
        self.filename = filename
        self.wb = px.load_workbook(filename=filename, read_only=False)
        self.sheetnames = self.wb.sheetnames

    def __getitem__(self, key):
        assert key in self.sheetnames
        return self.get_sheet(key)

    @staticmethod
    def _get_num_index(letter_idx):
        """文字インデックスから数字インデックスを計算(one-based index)"""
        assert isinstance(letter_idx, str)
        assert re.match(r'[A-Z]+$', letter_idx)
        idx = 0
        for i, l in enumerate(letter_idx[::-1]):
            w = 26 ** i
            x = ExcelWrapper.LETTERS.index(l) + 1
            idx += x * w
        return idx

    @staticmethod
    def _get_letter_index(num_idx, log=False):
        """数字インデックスから文字インデックスを計算(one-based index)"""
        assert isinstance(num_idx, int), "num_index: {}".format(num_idx)
        assert num_idx > 0
        N = 26
        num_idx -= 1

        def search_digit(idx, radix):
            i = 1
            offset = 0
            while True:
                pow_ = radix ** i
                if idx < pow_ + offset:
                    return i, offset
                i += 1
                offset += pow_

        def conv_radix(dec, radix):
            """A-Zを使わない基数変換"""
            if dec == 0: return [0]
            rs = []
            while dec > 0:
                q = dec / radix
                r = dec % radix
                rs.append(r)
                dec = q
            rs.reverse()
            return rs # 素因数分解の結果

        def split_(itr, n):
            l = len(itr)
            size = l / n + (l % n> 0)
            return [itr[i:i+size] for i in xrange(0, l, size)]

        def print_log(msg):
            if log:
                print msg

        print_log('num_idx - 1          : {}'.format(num_idx))

        digit, offset = search_digit(num_idx, N) # N進数において何桁かを求める
        print_log('digit_class({}base)  : {}'.format(N, digit))
        print_log('idx_offset           : {}'.format(offset))

        pf_idx = num_idx - offset                # digit桁の数列のインデックス(10進数)
        print_log('idx_in_digit_class   : {}'.format(pf_idx))

        pfs = conv_radix(pf_idx, N)              # インデックスを素因数分解
        print_log('idx({}base)          : {}'.format(N, pfs))

        while len(pfs) < digit: pfs.insert(0, 0) # digitに合わせて桁を調整

        print_log('zerofilled_idx       : {}'.format(pfs))

        ret = ''.join(ExcelWrapper.LETTERS[i] for i in pfs)
        print_log('Result:')
        print_log('num: {}'.format(num_idx + 1))
        print_log('str: {}'.format(ret))
        return ret

    @staticmethod
    def _letter_idx_gen(letter_idx_range):
        """レターインデックスのジェネレータ"""
        min_i, max_i = letter_idx_range
        assert re.match(r'[A-Z]+', min_i)
        assert re.match(r'[A-Z]+', max_i)
        f = ExcelWrapper._get_num_index
        for i in xrange(f(min_i), f(max_i) + 1):
            yield ExcelWrapper._get_letter_index(i)

    def get_sheet(self, sheetname):
        """Excelシートのオブジェクトを返す

        :param sheetname: str
            sheetname in workbook

        :return sheet: ExcelWrapper.Sheet
            Sheet object

        """

        return self.Sheet(self.wb, sheetname)

    class Sheet(object):

        def __init__(self, wb, sheetname):
            self.wb = wb
            self.ws = wb[sheetname]

        def _select(self, key, log=False):
            """min_col, min_row, max_col, max_row"""
            key_len = len(key)
            assert key_len in (2, 4)
            if key_len == 2:
                if log: print "{}{}を読み込み中です...".format(*key)
                return self.ws['{}{}'.format(*key)]
            elif key_len == 4:
                if log: print "{}{}:{}{}を読み込み中です...".format(*key)
                return self.ws['{}{}:{}{}'.format(*key)]

        def pickup_cell(self, coord, log=False):
            """セルを読み込む

            :param coord : tuple
                セルの座標
                ('A', 1): 'A1'のセル

            """
            return self._select(coord).value

        def iter_cell(self, coords, log=False):
            """セルをイテレーション

            :param coords : iterable of tuple
                セルの座標(tuple)のiterable

            """

            assert hasattr(coords, '__iter__')

            for coord in coords:
                yield self.pickup_cell(coord)

        def get_col(self, col, row_range=(1, None), iter_cell=False, log=False):
            """1列だけ読み込む

            :param col : str
                読み込む列の文字列

            :param row_range : iterable of ints, default: (1, None)
                読み込む行の範囲
                [開始行, 終了行]

            :param iter_cell : bool, default: False
                列の表現方法
                False: リスト
                True : セルのイテレータ

            :param log : bool, default: False
                ログを出力するかどうか

            :return column : list or generator
                iter_cellがTrueの場合はセルの値のイテレータ

            """

            r1, r2 = row_range
            if r2 is None:
                r2 = self.ws.max_row
            if iter_cell:
                return (self._select((col, i), log).value for i in xrange(r1, r2 + 1))
            else:
                return [cell[0].value for cell in self._select((col, r1, col, r2), log)]

        def get_row(self, row, col_range=('A', None), iter_cell=False, log=False):
            """1行だけ読み込む

            :param row : int
                読み込む行のint

            :param col_range : iterable of str
                読み込む列の範囲
                [開始列, 終了列]

            :param iter_cell : bool, default: False
                行の表現方法
                False: リスト
                True : セルのイテレータ

            :param log : bool, default: False
                ログを出力するかどうか

            :return column : list or generator
                iter_cellがTrueの場合はセルの値のイテレータ
            """

            c1, c2 = col_range
            if c2 is None:
                c2 = ExcelWrapper._get_letter_index(self.ws.max_column)
            if iter_cell:
                c1, c2 = (ExcelWrapper._get_num_index(c) for c in (c1, c2))
                return (self._select((ExcelWrapper._get_letter_index(i), row), log).value for i in xrange(c1, c2+1))
            else:
                return [cell.value for cell in self._select((c1, row, c2, row), log)[0]]

        def iter_cols(self, cols, row_range=(1, None), iter_cell=False, log=False, mode='line'):
            """指定した列のジェネレータ

            1列のみ読み込む場合はget_single_col()の使用を推奨

            :param cols : iterable
                読み込むいくつかの列
                modeにより振る舞いを指定

            :param row_range : tuple, default: (1, None)
                読み込む行の範囲
                (開始行, 終了行)
                終了行にNoneを指定するとデータのある最後の行まで読み込む

            :param iter_cell : bool, default: False
                列の表現方法
                False: リスト
                True : セルのイテレータ

            :param log : bool, default: False
                ログを出力するかどうか

            :param mode : str default: 'line'
                列の指定モード
                'line': colsを個別に指定
                    (A, C) -> 'A', 'C'
                'rect': colsを範囲で指定
                    (A, C) -> 'A', 'B', 'C'

            :return column_iter : iterator of list or iterator of iterator
                指定した列のイテレータ
                iter_cellがTrueの場合はセルの値のイテレータのイテレータ

            """

            assert hasattr(cols, '__iter__')
            assert isinstance(row_range, tuple)

            def is_collect_col(c):
                if not isinstance(c, str):
                    return False
                if re.match(r'[A-Z]+$', c) is not None:
                    return True
                return False

            assert False not in map(is_collect_col, cols)
            assert mode in ('line', 'rect')

            if mode == 'line':
                for col in cols:
                    yield self.get_col(col, row_range, iter_cell, log)
            elif mode == 'rect':
                for col in ExcelWrapper._letter_idx_gen(cols):
                    yield self.get_col(col, row_range, iter_cell, log)

        def iter_rows(self, rows, col_range=('A', None), iter_cell=False, log=False, mode='line'):
            """指定した行のジェネレータを生成

            1行のみ読み込む場合はget_single_row()の使用を推奨

            :param rows : iterable of int or str
                読み込むいくつかの行
                (1, 2)         : 1, 2行を読み込む

            :param col_range : tuple, default: ('A', None)
                読み込む列の範囲
                (開始列, 終了列)
                終了列にNoneを指定するとデータのある最後の列まで読み込む

            :param iter_cell : bool, default: False
                行の表現方法
                False: リスト
                True : セルのイテレータ

            :param log : bool, default: False
                ログを出力するかどうか

            :param mode : str default: 'line'
                行の指定モード
                'line': rowsを個別に指定
                    (1, 3) -> 1, 3
                'rect': rowsを範囲で指定
                    (1, 3) -> 1, 2, 3

            :return row_iter : iterator of list or iterator of iterator
                指定した行のイテレータ
                iter_cellがTrueの場合はセルの値のイテレータのイテレータ

            """

            assert hasattr(rows, '__iter__')
            assert isinstance(col_range, tuple)

            def is_correct_row(r):
                if isinstance(r, int):
                    return True
                return False

            assert False not in map(is_correct_row, rows)
            assert mode in ('line', 'rect')

            if mode == 'line':
                for row in rows:
                    yield self.get_row(row, col_range, iter_cell, log)
            elif mode == 'rect':
                for row in xrange(rows[0], rows[1] + 1):
                    yield self.get_row(row, col_range, iter_cell, log)

        def iter_part_col(self, col, length, row_range=(1, None), rest=False, log=False):
            """1列を分割してイテレーション

            列colを長さlengthごとにで分割したリストのジェネレータ

            :param col : str
                読み込む列

            :param row_range : tuple
                読み込む行の範囲
                (開始行, 終了行)

            :param length : int
                列を分割する長さ

            :param rest : bool, default: False
                戻り値にlen() < lengthのリストを含めるか

            :return part_col_iter : iterator of list
                分割された列をリストとしたイテレータ

            """

            assert isinstance(row_range, tuple)
            assert len(row_range) == 2
            assert length > 0

            r1 = row_range[0]
            limit_row = row_range[1] if row_range[1] is not None else self.ws.max_row
            r2 = r1 + length - 1
            is_last = False
            while True:
                part_col = self.get_col(col, (r1, r2), iter_cell=False, log=log)
                if not part_col: break
                #if log: print "part_col_len:", len(part_col)
                yield part_col
                r1 += length
                r2 += length
                if r2 > limit_row:
                    if is_last: break
                    if rest:
                        r2 -= r2 - limit_row
                        is_last = True
                        continue
                    else:
                        break

if __name__ == '__main__':
    from util import timecounter
    @timecounter
    def test_get_letter():
        import random
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        itr =1000
        max_digit = 3
        correct_n = 0
        incorrects = []
        for i in xrange(itr):
            r = random.randint(1, max_digit)
            t = "".join(random.choice(letters) for i in xrange(r))
            n = ExcelWrapper._get_num_index(t)
            l = ExcelWrapper._get_letter_index(n, True)
            print
            if t == l:
                correct_n += 1
            else:
                incorrects.append("{} : {}".format(n, l))
        print "Result: {} / {}".format(correct_n, itr)
        if not incorrects:
            print 'OK'
        else:
            for i in incorrects: print i

    def test_next_letter(l):
        print ExcelWrapper._get_next_letter(l)

    def main():
        filepath = r'E:\read_test.xlsx'
        wb = ExcelWrapper(filepath)
        ws = wb.get_sheet('Sheet1')
        print list(ws.iter_cell(coords=(('A', i) for i in xrange(1, 10, 2)), log=True))

    test_get_letter()
